# import openpyxl module
import time

import openpyxl
import sys
import os
import subprocess
import xlsxwriter

from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

workbook = xlsxwriter.Workbook('comments.xlsx')
worksheet = workbook.add_worksheet()

driver = webdriver.Edge()

# Declare the variables
global path
global wb_obj
global sheet_obj
global row
global column
columnNumbers = list()
global fromRow
global toRow

print("length of arguments should be 4: ")
print(len(sys.argv))
if len(sys.argv) == 4:
    # Check if the given excel file exists
    if os.path.exists(sys.argv[1]):
        path = sys.argv[1]
        # Load the Excel file
        wb_obj = openpyxl.load_workbook(path)
        # Get the active sheet
        sheet_obj = wb_obj.active
        # Getting the value of maximum rows and column
        row = sheet_obj.max_row
        column = sheet_obj.max_column
    else:
        print(sys.argv[1] + " not found")

    if len(sys.argv[2]) > 0:
        # Assign column number
        print("Columns: " + sys.argv[2])
        columnNumbers = sys.argv[2].split(',')
    else:
        print("Invalid column number")
    if len(sys.argv[3]) > 0:
        # Assign row number
        fromRow = int(sys.argv[3])
    else:
        print("Invalid fromRow number")

    for i in range(fromRow, row):
        cell_obj = sheet_obj.cell(row=i, column=int(columnNumbers[0]))
        cell_disconnected = sheet_obj.cell(row=i, column=int(columnNumbers[1]))
        print("On row " + str(i))
        address = cell_obj.value
        disconnected = cell_disconnected.value
        print(address)
        print(disconnected)
        if disconnected == "YES" and address is not None:
            res = subprocess.call(['ping', '-n', '3', address])
            if res == 0:
                print("ping to", address, "OK")
                try:
                    driver.get("http://" + address + "/")
                    # time.sleep(10)
                    wait = WebDriverWait(driver, 15)
                    # enter = input("Press Enter to Continue \n")
                    # print(enter)
                    if driver.title == "AXIS":
                        print("Axis Camera detected. Continuing to the next camera")
                        worksheet.write('A'+str(i), 'Connected')
                        # do something with that
                        continue
                    elif driver.title == "Login":
                        wait.until(EC.element_to_be_clickable((By.ID, "login_logo")))
                        driver.find_element(By.CLASS_NAME, "login-logo-dhtech")
                        print("Dahua Camera detected. Continuing to the next camera")
                        worksheet.write('A' + str(i), 'Connected')
                    elif driver.title == address:
                        wait.until(EC.element_to_be_clickable((By.ID, "logoImg")))
                        print("UNV Camera detected. Continuing to the next camera")
                        worksheet.write('A' + str(i), 'Connected')
                    # wait.until(EC.alert_is_present())
                    print("Waiting for element to be clickable")
                    print("Element clicked")
                except WebDriverException:
                    print("Could not reach site. Continuing to next camera")
                    worksheet.write('A' + str(i), 'Try again manually')
                    continue
            elif res == 2:
                print("no response from", address)
                worksheet.write('A' + str(i), 'No Response from ping')
            else:
                print("ping to", address, "failed!")
                worksheet.write('A' + str(i), 'Ping to camera failed')

            # cont = input("Continue to next camera? y/n \n")
            # if cont != "y":
            #     driver.close()
            #     break
            try:
                driver.close()
            except WebDriverException:
                print("No Window Open")
                driver = webdriver.Edge()
                # do something with that
                continue
        else:
            print("Connected")
            worksheet.write('A' + str(i), 'Connected')
    try:
        driver.close()
    except WebDriverException:
        print("No Window Open")
        # do something with that
    # Finally, close the Excel file
    # via the close() method.
    workbook.close()
else:
    print("Usage: ipSearch.py <filePath> <columnNumber> <fromRow> <toRow>")
