# Python program to read an excel file 
  
# import openpyxl module 
import openpyxl
import sys
import os
import subprocess

# Declare the variables
global path
global wb_obj
global sheet_obj
global row
global column
global columnNumber
global fromRow
global toRow

print("length of arguments should be 2 or 5: ")
print(len(sys.argv))
if len(sys.argv) == 4:
    # Check if the given excel file exists
    if os.path.exists(sys.argv[1]):
        path = sys.argv[1]
        # Load the Excel file
        wb_obj = openpyxl.load_workbook(path)
        # Get the active sheet
        sheet_obj = wb_obj.active
        # Getting the value of maximum rows and column
        row = sheet_obj.max_row
        column = sheet_obj.max_column
    else:
        print(sys.argv[1] + " not found")

    if int(sys.argv[2]) > 0:
        # Assign column number
        columnNumber = int(sys.argv[2])
    else:
        print("Invalid column number")
    if int(sys.argv[3]) > 0:
        # Assign row number
        fromRow = int(sys.argv[3])
    else:
        print("Invalid fromRow number")

    for i in range(fromRow, row): 
        cell_obj = sheet_obj.cell(row = i, column = columnNumber)
        address = cell_obj.value
        res = subprocess.call(['ping', '-n', '3', address])
        if res == 0:
            print( "ping to", address, "OK")
        elif res == 2:
            print ("no response from", address)
        else:
            print ("ping to", address, "failed!")
else:
    print("Usage: ipSearch.py <filePath> <columnNumber> <fromRow> <toRow>")