
# Python program to read an excel file 
  
# import openpyxl module 
import openpyxl
import webbrowser
import sys
import os
import subprocess

# File path to Edge
edge_path="C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
# Register Edge with the webbrowser
webbrowser.register('edge', None, webbrowser.BackgroundBrowser(edge_path))

# Declare the variables
global path
global wb_obj
global sheet_obj
global row
global column
global columnNumber
global fromRow
global toRow

print("length of arguments should be 2 or 5: ")
print(len(sys.argv))
if len(sys.argv) == 4:
    # Check if the given excel file exists
    if os.path.exists(sys.argv[1]):
        path = sys.argv[1]
        # Load the Excel file
        wb_obj = openpyxl.load_workbook(path)
        # Get the active sheet
        sheet_obj = wb_obj.active
        # Getting the value of maximum rows and column
        row = sheet_obj.max_row
        column = sheet_obj.max_column
    else:
        print(sys.argv[1] + " not found")

    if int(sys.argv[2]) > 0:
        # Assign column number
        columnNumber = int(sys.argv[2])
    else:
        print("Invalid column number")
    if int(sys.argv[3]) > 0:
        # Assign row number
        fromRow = int(sys.argv[3])
    else:
        print("Invalid fromRow number")

    for i in range(fromRow, row): 
        cell_obj = sheet_obj.cell(row = i, column = columnNumber)
        webbrowser.get('edge').open_new_tab(cell_obj.value)
        address = cell_obj.value
        res = subprocess.call(['ping', '-c', '3', address])
    if res == 0:
        print("ping to", address, "OK")
    elif res == 2:
        print("no response from", address)
    else:
        print("ping to", address, "failed!")

elif len(sys.argv) == 5:
    # Check if the given excel file exists
    if os.path.exists(sys.argv[1]):
        path = sys.argv[1]
        # Load the Excel file
        wb_obj = openpyxl.load_workbook(path)
        # Get the active sheet
        sheet_obj = wb_obj.active
        # Getting the value of maximum rows and column
        row = sheet_obj.max_row
        column = sheet_obj.max_column
    else:
        print(sys.argv[1] + " not found")
    if int(sys.argv[2]) > 0:
        # Assign column number
        columnNumber = int(sys.argv[2])
    else:
        print("Invalid column number")
    if int(sys.argv[3]) > 0:
        # Assign row number
        fromRow = int(sys.argv[3])
    else:
        print("Invalid fromRow number")

    if int(sys.argv[4]) > fromRow:
        if (int(sys.argv[4]) - fromRow) > 20:
            toRow = fromRow + 20
            print("Too many tabs to open. Opening values from row: " + str(fromRow) + " to row: " + str(toRow))
        else:
            toRow = int(sys.argv[4])
        for i in range(fromRow, toRow): 
            cell_obj = sheet_obj.cell(row = i, column = columnNumber) 
            webbrowser.get('edge').open_new_tab(cell_obj.value)
    else:
        print("Invalid toRow number")
else:
    print("Usage: ipSearch.py <filePath> <columnNumber> <fromRow> <toRow>")
