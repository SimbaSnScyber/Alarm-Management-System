from openpyxl import load_workbook
from yattag import Doc, indent

wb = load_workbook("EBM Events.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = ''
xml_schema = ''

doc.asis(xml_header)
doc.asis(xml_schema)

with tag('XML'):
    with tag('AllEvents'):
        # Use ws.max_row for all rows
        for row in ws.iter_rows(min_row=137, max_row=226, min_col=1, max_col=2):
            row = [cell.value for cell in row]
            print("Row 0 is: " + row[0])
            print("Row 1 is: " + row[1])
            with tag("Events"):
                with tag("EventName"):
                    text(row[0])
                with tag("EventDescription"):
                    text(row[1])
                with tag("EventType"):
                    text("ALARM")
                with tag("IsSoundEnabled"):
                    text("false")
                with tag("IsNetworkDisabled"):
                    text("false")
                with tag("IsProtocolDisabled"):
                    text("false")
                with tag("IsWindowsLogEnabled"):
                    text("false")

    # <Events>
    #   <EventName>IM_P_VIBRATION_SENSOR_T_BYPASS</EventName>
    #   <EventDescription>Parked Vibration Sensor 3 Bypass (IM)</EventDescription>
    #   <EventType>INFORMATION</EventType>
    #   <IsSoundEnabled>false</IsSoundEnabled>
    #   <IsNetworkDisabled>false</IsNetworkDisabled>
    #   <IsProtocolDisabled>false</IsProtocolDisabled>
    #   <IsWindowsLogEnabled>false</IsWindowsLogEnabled>
    # </Events>
result = indent(doc.getvalue())
print(result)

with open("events.xml", "w") as f:  #give path where you want to create
    f.write(result)
